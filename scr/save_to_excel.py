from openpyxl import load_workbook, Workbook
import os


class SaveToExcel:
    def __init__(self,overlap=False):
        file_name = './output/boss_crawl_output.xlsx'
        if not os.path.exists('./output'): os.mkdir('./output')
        if not overlap:
            counter = 1
            while os.path.exists(file_name):
                file_name = f'./output/boss_crawl_output_{counter}.xlsx'
                counter += 1
        elif os.path.exists(file_name): os.remove(file_name)

        self.file_path = file_name

        wb = Workbook()
        ws = wb.active
        row = ws.max_row

        ws.cell(row=row, column=1, value='职位')
        ws.cell(row=row, column=2, value='薪资')
        ws.cell(row=row, column=3, value='工作岗位标签')
        ws.cell(row=row, column=4, value='公司名')
        ws.cell(row=row, column=5, value='公司标签')
        ws.cell(row=row, column=6, value='工作详情页')
        ws.cell(row=row, column=7, value='职位沟通链接')

        wb.save(self.file_path)
        wb.close()

    def save(self,job_name,job_salary,job_tags,job_company,job_company_tags,job_content,job_link):
        wb = load_workbook(self.file_path)
        ws = wb.active
        row = ws.max_row + 1

        ws.cell(row=row, column=1, value=job_name)
        ws.cell(row=row, column=2, value=job_salary)
        ws.cell(row=row, column=3, value=job_tags)
        ws.cell(row=row, column=4, value=job_company)
        ws.cell(row=row, column=5, value=job_company_tags)
        ws.cell(row=row, column=6, value=job_content)
        ws.cell(row=row, column=7, value=job_link)

        wb.save(self.file_path)
        wb.close()
